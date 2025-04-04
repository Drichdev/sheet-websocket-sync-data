# python 3.10
from flask import Flask, render_template, request, redirect, url_for, session
from flask_socketio import SocketIO, emit
import openpyxl
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret!'
socketio = SocketIO(app)

EXCEL_FILE = 'data.xlsx'
ADMIN_EMAIL = 'admin@example.com'
ADMIN_PASSWORD = 'admin'

def init_excel():
    """Créer le fichier Excel avec les en-têtes s'il n'existe pas."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Définition des en-têtes
        ws.append(["Nom", "Prénom", "Email"])
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html')

@socketio.on('submit_data')
def handle_submit(data):
    print('Données reçues :', data)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([data.get("nom"), data.get("prenom"), data.get("email")])
    wb.save(EXCEL_FILE)
    # Envoyer une confirmation à tous les clients connectés
    emit('data_saved', data, broadcast=True)


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            session['admin_email'] = email
            return redirect(url_for('admin'))
        
        else:
            error = "Identifiants invalides."
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('login'))

@app.route('/admin')
def admin():
    if not session.get('admin_logged_in'):
        return redirect(url_for('login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data.append((idx, row))
    return render_template('admin.html', headers=headers, data=data, admin_email=session.get('admin_email'))


@app.route('/delete', methods=['POST'])
def delete_row():
    if not session.get('admin_logged_in'):
        return redirect(url_for('login'))
    
    try:
        # Le formulaire envoie l'indice de la ligne à supprimer
        row_index = int(request.form.get('row_index'))
    except (TypeError, ValueError):
        return redirect(url_for('admin'))
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    # row_index correspond à la ligne Excel
    ws.delete_rows(row_index)
    wb.save(EXCEL_FILE)
    return redirect(url_for('admin'))

if __name__ == '__main__':
    init_excel()
    socketio.run(app, debug=True)
